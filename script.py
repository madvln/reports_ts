import os
import pdfkit
from jinja2 import Environment, FileSystemLoader
from openpyxl import load_workbook
from concurrent.futures import ThreadPoolExecutor

# Путь к папкам
templates_dir = "templates"
data_dir = "data"
output_dir = "pdf"
path_to_wkhtmltopdf = (
    r"C:\Users\tarakanchikoves\source\reports_ts\wkhtmltox\bin\wkhtmltopdf.exe"
)
config = pdfkit.configuration(wkhtmltopdf=path_to_wkhtmltopdf)


# # Функция для извлечения города из имени файла
# def get_city_from_filename(filename):
#     return filename.replace("data_", "").replace(".xlsx", "")


# Функция для загрузки данных из Excel
def load_excel_data(file_path):
    wb = load_workbook(file_path)
    sheet = wb.active
    data = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        data.append(
            {
                "key_1": row[0],  # Предположим, что столбцы в Excel следующие
                "key_2": row[1],
                "key_4": row[3],
            }
        )
    return data


# Функция для рендеринга HTML из шаблона
def render_html(template_name, data):
    env = Environment(loader=FileSystemLoader(templates_dir))
    template = env.get_template(template_name)
    return template.render(data)


# Функция для конвертации HTML в PDF и сохранения в папку
def convert_html_to_pdf(html_content, output_path):
    pdfkit.from_string(html_content, output_path, configuration=config)


# Функция для генерации PDF для одного города
def generate_pdfs_for_city(city_name):
    # Загружаем данные из Excel
    data_file = os.path.join(data_dir, f"data_{city_name}.xlsx")
    data = load_excel_data(data_file)

    # Создаем папку для города, если она не существует
    city_output_dir = os.path.join(output_dir, city_name)
    os.makedirs(city_output_dir, exist_ok=True)

    # Загружаем соответствующий HTML-шаблон
    template_name = f"{city_name}.htm"

    # Генерация PDF для каждой строки данных
    for i, row_data in enumerate(data):
        # Рендерим HTML с данными
        html_content = render_html(template_name, row_data)
        # print(row_data["key_4"].split(",")[1])
        # Путь к PDF-файлу
        pdf_filename = f"{row_data["key_4"].replace(",","")}.pdf"
        pdf_path = os.path.join(city_output_dir, row_data["key_4"].split(",")[1])
        os.makedirs(pdf_path, exist_ok=True)
        pdf_path = os.path.join(pdf_path, pdf_filename)
        # Конвертируем HTML в PDF
        convert_html_to_pdf(html_content, pdf_path)
        print(f"PDF для {city_name} создан: {pdf_path}")


# Основная логика для распараллеливания
def generate_pdfs():
    # Сканируем папку templates для всех файлов .htm (городов)
    cities = [
        filename.split(".")[0]
        for filename in os.listdir("templates")
        if filename.endswith(".htm")
    ]

    # Использование ThreadPoolExecutor для параллельной обработки
    with ThreadPoolExecutor(max_workers=4) as executor:
        executor.map(generate_pdfs_for_city, cities)


if __name__ == "__main__":
    generate_pdfs()
